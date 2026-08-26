#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
import_f259_source.py — กู้ค่า 4 คอลัมน์จาก F-259 เข้าตาราง lg_req_f259_source (migration 052)

อ่าน:  lexguard/local-data/F-259_2569_R1.xlsx  (gitignore แล้ว — ห้าม commit)
เขียน: lg_req_f259_source เท่านั้น

═══ ปลอดภัยต่อข้อมูลเดิม ═══
· INSERT ลงตารางใหม่อย่างเดียว — ไม่มี UPDATE/DELETE กับตารางใดทั้งสิ้น
· ไม่แตะ lg_requirements แม้แต่คอลัมน์เดียว (กติกา 9.2)
· รันซ้ำได้ — ล้างเฉพาะแถวที่สคริปต์นี้เคยใส่เอง (imported_by = ค่าเดียวกัน) ก่อนใส่ใหม่

การจับคู่กับข้อกำหนดในทะเบียน:
  exact   ข้อความหลัง normalize ตรงกันทั้งหมด
  partial ข้อความ 60 ตัวอักษรแรกอยู่ในกันและกัน (ข้อมูลถูกแก้คำเล็กน้อยหลังนำเข้า)
  unmatched จับคู่ไม่ได้ — ยังเก็บไว้ (requirement_id = null) ให้คนมาไล่เอง ไม่ทิ้งเงียบ

หมวด CCS (FG - CCS) ถูกข้าม เพราะหมวด CC ถูกถอดออกจากแอปตั้งแต่ P15
"""
import openpyxl, datetime, re, json, os, sys, urllib.request, urllib.parse

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(ROOT, 'local-data', 'F-259_2569_R1.xlsx')
TAG  = 'scripts/import_f259_source.py'

env = {}
for line in open(os.path.join(ROOT, '.env'), encoding='utf-8'):
    if '=' in line and not line.strip().startswith('#'):
        k, v = line.strip().split('=', 1); env[k] = v
U, K = env['VITE_SUPABASE_URL'], env['VITE_SUPABASE_ANON_KEY']
H = {'apikey': K, 'Authorization': 'Bearer ' + K, 'content-type': 'application/json'}

def api(path, method='GET', body=None, prefer=None):
    h = dict(H)
    if prefer: h['Prefer'] = prefer
    data = json.dumps(body).encode() if body is not None else None
    req = urllib.request.Request(f"{U}/rest/v1/{path}", data=data, headers=h, method=method)
    with urllib.request.urlopen(req) as r:
        t = r.read().decode()
        return json.loads(t) if t else []

SHEETS = [('LA','LA-บริหารจัดการความปลอดภัย'), ('LB','LB-ไฟฟ้าและพลังงาน'), ('LC','LC-อัคคีภัย (OK)'),
          ('LD','LD-ร้อน แสง เสียง สภาพแวดล้อม'), ('LE','LE-ก่อสร้าง ลิฟต์  เครื่องจักร '),
          ('LG','LG-คณะกรรมการสวัสดิการ.'), ('LF','LF- Service')]
CODE_RE = re.compile(r'^[A-Za-z]{1,3}[- ]?\d{1,3}')

def s(v):
    if v is None: return ''
    if isinstance(v, datetime.datetime): return v.strftime('%d/%m/%Y')
    return str(v).strip()

def findcol(ws, kw):
    for r in (5, 6, 7):
        for c in range(1, ws.max_column + 1):
            if kw in s(ws.cell(r, c).value): return c

norm = lambda x: re.sub(r'\s+', '', re.sub(r'[^฀-๿a-zA-Z0-9]', '', str(x or '')))[:120]

def main():
    if not os.path.exists(XLSX):
        sys.exit(f'ไม่พบไฟล์ {XLSX} — ไฟล์นี้อยู่เฉพาะบนเครื่องผู้ดูแล')
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    laws = api('lg_laws?select=id,cat,code&limit=500')
    lawkey = {(l['cat'], l['code']): l['id'] for l in laws}
    reqs = api('lg_requirements?select=id,law_id,text&limit=1000')
    byLaw = {}
    for r in reqs: byLaw.setdefault(r['law_id'], []).append(r)

    rows, stat = [], {'exact': 0, 'partial': 0, 'unmatched': 0}
    for cat, sheet in SHEETS:
        ws = wb[sheet]
        col = {k: findcol(ws, k) for k in
               ['ชื่อกฎหมาย', 'สรุปสาระสำคัญ', 'ผู้รับผิดชอบ', 'ความถี่', 'การรายงานผล', 'เอกสารที่เกี่ยวข้อง', 'หมายเหตุ']}
        cur = None
        for r in range(8, ws.max_row + 1):
            code, name = s(ws.cell(r, 1).value), s(ws.cell(r, col['ชื่อกฎหมาย']).value)
            if code and CODE_RE.match(code) and name: cur = code
            txt = s(ws.cell(r, col['สรุปสาระสำคัญ']).value)
            if not txt or not cur: continue
            vals = {k: (s(ws.cell(r, col[k]).value) if col[k] else '')
                    for k in ['ผู้รับผิดชอบ', 'ความถี่', 'การรายงานผล', 'เอกสารที่เกี่ยวข้อง', 'หมายเหตุ']}
            # แถวที่ไม่มีอะไรให้เติมเลย ไม่ต้องเก็บ — คิวจะบวมโดยไม่ได้อะไร
            if not any(vals[k] for k in ['ผู้รับผิดชอบ', 'ความถี่', 'การรายงานผล', 'เอกสารที่เกี่ยวข้อง']): continue

            law_id = lawkey.get((cat, cur))
            rid, kind = None, 'unmatched'
            if law_id:
                n = norm(txt)
                cand = byLaw.get(law_id, [])
                hit = next((c for c in cand if norm(c['text']) == n), None)
                if hit: rid, kind = hit['id'], 'exact'
                else:
                    hit = next((c for c in cand if n[:60] and (n[:60] in norm(c['text']) or norm(c['text'])[:60] in n)), None)
                    if hit: rid, kind = hit['id'], 'partial'
            stat[kind] += 1
            rows.append({
                'requirement_id': rid, 'law_id': law_id,
                'cat': cat, 'law_code': cur, 'source_sheet': sheet, 'source_row': r,
                'req_excerpt': txt[:200],
                'responsible': vals['ผู้รับผิดชอบ'] or None,
                'frequency':   vals['ความถี่'] or None,
                'documents':   vals['เอกสารที่เกี่ยวข้อง'] or None,
                'report_to':   vals['การรายงานผล'] or None,
                'remark':      vals['หมายเหตุ'] or None,
                'match_kind': kind, 'imported_by': TAG,
            })

    print(f"อ่านได้ {len(rows)} แถว — ตรงเป๊ะ {stat['exact']} · บางส่วน {stat['partial']} · จับคู่ไม่ได้ {stat['unmatched']}")
    if '--dry-run' in sys.argv:
        print('โหมดทดลอง — ไม่เขียนอะไรลงฐาน'); return

    # ล้างเฉพาะแถวที่สคริปต์นี้เคยใส่ (รันซ้ำได้ ไม่แตะข้อมูลของใครอื่น)
    api(f"lg_req_f259_source?imported_by=eq.{urllib.parse.quote(TAG)}", method='DELETE')
    for i in range(0, len(rows), 200):
        api('lg_req_f259_source', method='POST', body=rows[i:i+200], prefer='return=minimal')
        print(f"  เขียนแล้ว {min(i+200, len(rows))}/{len(rows)}")
    print('เสร็จ — lg_requirements ไม่ถูกแตะแม้แต่แถวเดียว')

if __name__ == '__main__':
    main()
